#!/usr/bin/env python3
"""Build data/dict.json — lookup dictionary for the in-app "My words" feature.

Covers the ~31k frequency-ranked lemmas from the NGSL source corpus that are
NOT already in data/words.json, so textbook words the user adds by hand can
get their Japanese definition auto-filled. Definitions come from ejdict-hand
(public domain), trimmed to the first 3 sense segments.

Usage:
  python3 tools/build_dict.py <ngsl.xlsx> <ejdict-all.txt> <words.json> > data/dict.json
"""
import json
import re
import sys

import openpyxl

WORD_RE = re.compile(r"[A-Za-z][A-Za-z'\-]*")
MAX_SENSES = 3


def load_ejdict(path):
    d = {}
    with open(path, encoding='utf-8') as f:
        for line in f:
            parts = line.rstrip('\n').split('\t')
            if len(parts) >= 2:
                for w in parts[0].split(','):
                    w = w.strip()
                    if w and w not in d:
                        d[w] = parts[1]
    return d


def main(xlsx_path, ejdict_path, words_path):
    ejdict = load_ejdict(ejdict_path)
    base = {e['w'] for e in json.load(open(words_path, encoding='utf-8'))}

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        lemma = row[0]
        if lemma is None:
            continue
        w = str(lemma).strip()
        key = w.lower()
        if key in base or key in out or not WORD_RE.fullmatch(w):
            continue
        ja = ejdict.get(w) or ejdict.get(key) or ejdict.get(w.capitalize())
        if not ja:
            continue
        out[key] = ' / '.join(ja.split(' / ')[:MAX_SENSES])

    print(f'dict entries: {len(out)}', file=sys.stderr)
    json.dump(out, sys.stdout, ensure_ascii=False, separators=(',', ':'))


if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2], sys.argv[3])

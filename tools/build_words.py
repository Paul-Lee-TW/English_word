#!/usr/bin/env python3
"""Build data/words.json — 3000 high-frequency English words with Japanese definitions.

Sources (both free to use):
  * NGSL 1.01 word list (CC BY 3.0, Browne/Culligan/Phillips)
      http://www.newgeneralservicelist.org/
      xlsx mirror: https://raw.githubusercontent.com/antdurrant/word.lists/master/data-raw/list_ngsl/NGSL%2B1.01%2Bwith%2BSFI.xlsx
  * ejdict-hand English-Japanese dictionary (Public Domain)
      https://github.com/kujirahand/EJDict  (src/a.txt ... z.txt concatenated)

Word selection: NGSL core 2801 + NGSL supplemental 47 (days/months/numbers),
then topped up to 3000 with the next highest-frequency lemmas (by SFI) from
the same corpus that have an ejdict entry.

Usage:
  python3 tools/build_words.py <ngsl.xlsx> <ejdict-all.txt> > data/words.json
"""
import json
import re
import sys

import openpyxl

# NGSL words missing from ejdict-hand, defined by hand.
MANUAL_JA = {
    'eventually': '結局,最終的に,ついに',
    'email': '〈U〉〈C〉電子メール,Eメール / …に電子メールを送る',
    'effectively': '効果的に,有効に / 事実上,実質的に',
    'anymore': '《否定文・疑問文で》今はもう,これ以上',
    'database': 'データベース',
    'initially': '初めに,最初は',
    'implementation': '実行,実施,履行',
    'fax': 'ファックス,ファクシミリ / …をファックスで送る',
    'jeans': 'ジーンズ,ジーパン',
    'website': 'ウェブサイト',
    'online': 'オンラインの,インターネットに接続された / オンラインで',
    'internet': '《the ~》インターネット',
    'smartphone': 'スマートフォン',
}
# NGSL entries not worth a junior-high deck (abbreviations / corpus noise).
SKIP = {'ph', 'en', 'oocyte'}

TARGET = 3000
WORD_RE = re.compile(r"[A-Za-z][A-Za-z'\-]*")


def load_ejdict(path):
    d = {}
    with open(path, encoding='utf-8') as f:
        for line in f:
            parts = line.rstrip('\n').split('\t')
            if len(parts) >= 2:
                for w in parts[0].split(','):
                    w = w.strip()
                    if w and w not in d:
                        d[w] = parts[1]
    return d


def main(xlsx_path, ejdict_path):
    ejdict = load_ejdict(ejdict_path)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row) + [None] * 4
        lemma, wordlist, sfi = row[0], row[1], row[3]
        if lemma is None or sfi is None:
            continue
        try:
            sfi = float(sfi)
        except (TypeError, ValueError):
            continue
        rows.append((str(lemma).strip(), wordlist, sfi))

    def lookup(w):
        if w.lower() in MANUAL_JA:
            return MANUAL_JA[w.lower()]
        for cand in (w, w.lower(), w.capitalize()):
            if cand in ejdict:
                return ejdict[cand]
        return None

    ngsl = [(w, s) for w, wl, s in rows if wl == '1 - NGSL']
    sup = [(w, s) for w, wl, s in rows if wl == '2 - Sup']
    rest = sorted([(w, s) for w, wl, s in rows if wl is None], key=lambda x: -x[1])

    words, seen, missing = [], set(), []

    def add(w):
        key = w.lower()
        if key in seen or key in SKIP or not WORD_RE.fullmatch(w):
            return
        ja = lookup(w)
        if ja is None:
            missing.append(w)
            return
        seen.add(key)
        words.append({'w': w.lower(), 'ja': ja})

    for w, _ in sorted(ngsl, key=lambda x: -x[1]):
        add(w)
    for w, _ in sorted(sup, key=lambda x: -x[1]):
        add(w)
    for w, _ in rest:
        if len(words) >= TARGET:
            break
        add(w)

    print(f'total: {len(words)}  missing: {missing}', file=sys.stderr)
    json.dump(words[:TARGET], sys.stdout, ensure_ascii=False, separators=(',', ':'))


if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2])
